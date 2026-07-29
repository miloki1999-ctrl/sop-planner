"""
Export service — spec section 24. One generic, reusable function
(`export_df_to_excel`) that produces professionally formatted Excel
(freeze header, autofilter, autofit columns, navy header style matching
the app's design system) so every page just builds a DataFrame and calls
this instead of reimplementing formatting.
"""
import io
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill(start_color="0B3D91", end_color="0B3D91", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
BODY_FONT = Font(name="Calibri", size=10)
THIN_BORDER = Border(*(Side(style="thin", color="D9D9D9"),) * 4)


def export_df_to_excel(df: pd.DataFrame, sheet_name: str = "Report", freeze_header: bool = True) -> bytes:
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=sheet_name[:31], index=False)
        ws = writer.sheets[sheet_name[:31]]

        for col_idx, col_name in enumerate(df.columns, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                cell.font = BODY_FONT
                cell.border = THIN_BORDER

        for col_idx, col_name in enumerate(df.columns, start=1):
            max_len = max([len(str(col_name))] + [len(str(v)) for v in df.iloc[:, col_idx - 1].astype(str)])
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 3, 10), 45)

        if freeze_header:
            ws.freeze_panes = "A2"
        if ws.max_row > 1:
            ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"

    return buf.getvalue()


def export_multi_sheet_excel(sheets: dict) -> bytes:
    """sheets = {'Sheet Name': DataFrame, ...}"""
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        for name, df in sheets.items():
            df.to_excel(writer, sheet_name=name[:31], index=False)
            ws = writer.sheets[name[:31]]
            for col_idx in range(1, len(df.columns) + 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
            for col_idx, col_name in enumerate(df.columns, start=1):
                max_len = max([len(str(col_name))] + [len(str(v)) for v in df.iloc[:, col_idx - 1].astype(str)]) if len(df) else len(str(col_name))
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 3, 10), 45)
            ws.freeze_panes = "A2"
            if ws.max_row > 1:
                ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    return buf.getvalue()
